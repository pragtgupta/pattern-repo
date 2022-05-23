import openpyxl


wb=openpyxl.load_workbook("/Users/pragtgupta/PycharmProjects/pythonProject/pac2/data.xlsx")
sh1=wb["employeetab"]

row=sh1.max_row
column=sh1.max_column

print(row)
print(column)

print("-----------")
for i in range(1, row+1):
    for j in range(1, column+1):
        print(sh1.cell(i, j).value)

#at the same time we can write down the data also


sh1.cell(row=6, column=1, value='lenin')
sh1.cell(row=6, column=2, value='mudugal')
sh1.cell(row=6, column=3, value='bhopal')
sh1.cell(row=6, column=4, value='88')


wb.save("Report.xlsx")
















