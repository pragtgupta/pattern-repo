import openpyxl

wb=openpyxl.load_workbook("/Users/pragtgupta/PycharmProjects/pythonProject/pac2/data.xlsx")  #wb is the object of workbook class

print(type(wb))
names = wb.sheetnames
print(names)

# print(wb.active.title)
#
sh1= wb["employeetab"]     #sh1 is the object
#
# print(type(sh1))

#first way---
print(wb["employeetab"]["A2"].value)  #pragti


#second sheet
print(wb["depttab"]["B2"].value)   #23



#second way----
print(sh1.cell(2,3).value) #ban  #row,column
print(sh1.cell(3,3).value) #del












