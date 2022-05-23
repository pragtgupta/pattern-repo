import sys
import os
import openpyxl

loc = ("/Users/pragtgupta/PycharmProjects/pythonProject/pac1/BookMyShowDatabase.xlsx")
wb=openpyxl.load_workbook("BookMyShowDatabase.xlsx")
login_sheet=wb['LoginData']            #getting the login sheet
ls_max_row=login_sheet.max_row         #max row in login sheet
ls_max_col=login_sheet.max_column      #max column in login sheet

Movie_Info_sheet=wb['MovieInfo']       #getting the movie sheet
mi_max_row=Movie_Info_sheet.max_row    #max row in movie sheet
mi_max_col=Movie_Info_sheet.max_column #max col in movie sheet

movie_rating_sheet=wb['MovieRating']   #getting movie rating sheet
mrs_max_row=movie_rating_sheet.max_row #max row in movieRating sheet
mrs_max_col=movie_rating_sheet.max_column #max col in movieRating sheet

movie_booking_sheet=wb['MovieBooking']   #get movie booking sheet
mb_max_row=movie_booking_sheet.max_row
mb_max_col=movie_booking_sheet.max_column




flag=0
mov_list=[]
movie_seat=0
movie_seat_row=0
user_seat_row=0
LoginData_User_row=0   #track the user row in loginsheet
mov_name_inMovRating=0 #track the column of movie name in Movie Rating sheet


print('****** Welcome to BookMyShow  *********')
print('Press 1 for login')
print('Press 2 for register a new account')
print('Press 3 for exit')

rolePath=int(input('Enter your choice:  '))
if rolePath==1:
    class login:
        print('******  Welcome to Login page  *********')
        user=input('Enter Username:  ')
        password=input('Enter password:  ')
        if user=='admin' and password=='admin':
            #inside admin
            print('*****  Welcome Admin  *****')
            print('Press 1 for add new movie info')
            print('Press 2 for edit new movie info')
            print('Press 3 for delete movie info')
            print('Press 4 for logout')
            admin_choice=int(input('Enter your choice:  '))
            if admin_choice==1:
                print('Welcome Admin, you are about to add a new movie')
                title=input('Enter Title:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1,column=1,value=title)
                genere=input('Enter Genere:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=2, value=genere)
                length=input('Enter Length:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=3, value=length)
                cast=input('Enter Casts:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=4, value=cast)
                director=input('Enter Director:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=5, value=director)
                rating=input('Enter Rating:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=6, value=rating)
                lang=input('Enter Language: ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=7, value=lang)
                s_time=input('Enter Time:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=8, value=s_time)
                seat=input('Enter Seat Capacity:  ')
                Movie_Info_sheet.cell(row=mi_max_row+1, column=9, value=seat)

                #Enter movie name on Movie Rating sheet
                mov_name_inMovRating=mrs_max_col+1  #track the coloumn of the movie in Movie Rating sheet
                movie_rating_sheet.cell(row=1,column=mov_name_inMovRating,value=title)

                # Enter movie name on Movie Rating sheet

                movie_booking_sheet.cell(row=1, column=mov_name_inMovRating, value=title)

                wb.save("BookMyShowDatabase.xlsx")
                print('You have successfully added movie details')


            elif admin_choice == 2:
                print('Welcome admin, you are about to edit a new movie')
                edit_mov = input('Enter the movie title that you want to edit:  ')
                for i in range(1, mi_max_row + 1):
                    if Movie_Info_sheet.cell(i, 1).value == edit_mov:
                        genere1 = input('Enter updated Genere:  ')
                        Movie_Info_sheet.cell(row=i, column=2, value=genere1)
                        length1 = input('Enter updated Length:  ')
                        Movie_Info_sheet.cell(row=i, column=3, value=length1)
                        cast1 = input('Enter updated Casts:  ')
                        Movie_Info_sheet.cell(row=i, column=4, value=cast1)
                        director1 = input('Enter updated  Director:  ')
                        Movie_Info_sheet.cell(row=i, column=5, value=director1)
                        rating1 = input('Enter updated Rating:  ')
                        Movie_Info_sheet.cell(row=i, column=6, value=rating1)
                        lang1 = input('Enter updated Language: ')
                        Movie_Info_sheet.cell(row=i, column=7, value=lang1)
                        s_time1 = input('Enter updated Time:  ')
                        Movie_Info_sheet.cell(row=i, column=8, value=s_time1)
                        seat1 = input('Enter updated Seat Capacity:  ')
                        Movie_Info_sheet.cell(row=i, column=9, value=seat1)
                        wb.save("BookMyShowDatabase.xlsx")
                        print('You have successfully updated movie details of %s' % (edit_mov))




                    elif admin_choice==3:
                       print('Welcome admin, you are about to delete a movie details')
                       del_mov=input('Enter the movie name that you want to delete:  ')
                       for i in range(1,mi_max_row+1):
                         if Movie_Info_sheet.cell(i,1).value==del_mov:
                               #movie row is deleted from movie info sheet
                                Movie_Info_sheet.delete_rows(idx=i,amount=1)
                print('Movie %s is deleted'%(del_mov))

                #Delete movie coloumn from Movie rating sheet
                for i in range(1,mrs_max_col+1):
                    if movie_rating_sheet.cell(1,i).value==del_mov:
                        print(movie_rating_sheet.cell(1,i).value)
                        movie_rating_sheet.delete_cols(idx=i, amount=1)

                #Delete movie column from movie booking sheet
                for i in range(1,mb_max_col+1):
                    if movie_booking_sheet.cell(1,i).value==del_mov:
                        print(movie_booking_sheet.cell(1,i).value)
                        movie_booking_sheet.delete_cols(idx=i, amount=1)


                wb.save("BookMyShowDatabase.xlsx")

            elif admin_choice==4:
                print('Thank you')
            else:
                print('You have entered a wrong choice')
                sys.exit()

        else:
            for i in range(2, ls_max_row + 1):
                if (login_sheet.cell(i, 1).value == user and login_sheet.cell(i, 2).value == password):
                    flag = 0
                    user_seat_row = i
                    print('*****  welcome %s  *****' % (user))
                    for i in range(2, mi_max_row + 1):
                        mov_list.append(Movie_Info_sheet.cell(i, 1).value)
                    for i in range(len(mov_list)):
                        print('Currently Running:   %s' % (mov_list[i]))

                    user_choice = input('Enter your movie name:  ')
                    for i in mov_list:
                        if i == user_choice:
                            print('Hi %s, see the Details of Movie:    %s  ' % (user, i))  # i==movie name
                            for j in range(2, mi_max_row + 1):
                                if i == Movie_Info_sheet.cell(j, 1).value:  # j=row no of the movie in excel
                                    for k in range(1, mi_max_col + 1):
                                        print('%s:  %s' % (
                                        Movie_Info_sheet.cell(1, k).value, Movie_Info_sheet.cell(j, k).value))
                                        # Return the movie seat from Movie Info page
                                        movie_seat = int(Movie_Info_sheet.cell(j, 9).value)
                                        movie_seat_row = j

                                    print()
                                    print('Press 1 to book a ticket')
                                    print('Press 2 to cancel ticket')
                                    print('Press 3 to give rating')
                                    user_choice1 = int(input('Enter your option:  '))
                                    if user_choice1 == 1:
                                        # Identify the cell in movie booking sheet and update value there
                                        for i in range(1, mb_max_row + 1):
                                            if movie_booking_sheet.cell(i, 1).value == user:
                                                ubp_user_row = i  # ubp_user_row -> row number of user in booking page

                                        print('Welcome to booking page')
                                        print('Remaining seats are:  %s' % (movie_seat))
                                        book_ticket = int(input('Enter the number of tickets you want to book:  '))
                                        if movie_seat >= book_ticket:
                                            print('You have successfully booked %s tickets' % (book_ticket))

                                            for i in range(1, mb_max_col + 1):
                                                if movie_booking_sheet.cell(1, i).value == user_choice:
                                                    movie_booking_sheet.cell(ubp_user_row, i).value = book_ticket
                                                    update_ticket = movie_seat - book_ticket
                                                    print('Remaining seats are: %s' % (update_ticket))
                                                    Movie_Info_sheet.cell(movie_seat_row, 9).value = update_ticket
                                                    wb.save("BookMyShowDatabase.xlsx")


                                        else:
                                            print('There are not enough tickets, please try again')


                                    elif user_choice1 == 2:
                                        print('Welcome to cancellation page')
                                        for i in range(1, mb_max_row + 1):
                                            if movie_booking_sheet.cell(i, 1).value == user:
                                                ubp_user_row = i  # ubp_user_row -> row number of user in booking page
                                        bt_Mov_col = 0

                                        for i in range(1, mb_max_col + 1):
                                            if movie_booking_sheet.cell(1, i).value == user_choice:
                                                bt_Mov_col = i
                                                booked_tickets = movie_booking_sheet.cell(ubp_user_row, i).value

                                        print('Previously you have booked %s tickets ' % (booked_tickets))
                                        cancel_tic = int(input('How many tickets you want to cancel:  '))
                                        if booked_tickets < cancel_tic:
                                            print('You does not have %s number od tickets' % (cancel_tic))
                                        else:
                                            # Updated ticket count movie info sheet
                                            Movie_Info_sheet.cell(movie_seat_row, 9).value += cancel_tic

                                            update_ticket = Movie_Info_sheet.cell(movie_seat_row, 9).value
                                            movie_booking_sheet.cell(ubp_user_row, bt_Mov_col).value -= cancel_tic
                                            wb.save("BookMyShowDatabase.xlsx")
                                            print('Cancellation successful, you have %s tickets left' % (
                                                movie_booking_sheet.cell(ubp_user_row, bt_Mov_col).value))
                                    elif user_choice1 == 3:
                                        print('Welcome to user feedback page')
                                        urp_user_row = 0
                                        feedback_ip_movie = user_choice
                                        # identify the current user row
                                        for i in range(1, mrs_max_row + 1):
                                            if movie_rating_sheet.cell(i, 1).value == user:
                                                urp_user_row = i

                                        for i in range(1, mrs_max_col + 1):
                                            if movie_rating_sheet.cell(1, i).value == feedback_ip_movie:
                                                if movie_rating_sheet.cell(urp_user_row, i).value != None:
                                                    print('Your feedback is already taken, earlier feedback:  %s' % (
                                                        movie_rating_sheet.cell(urp_user_row, i).value))
                                                else:
                                                    feedback_ip_rating = input('Enter your rating:  ')
                                                    movie_rating_sheet.cell(urp_user_row, i).value = feedback_ip_rating
                                                    print('Feedback taken, your current rating:  %s' % (
                                                        movie_rating_sheet.cell(urp_user_row, i).value))
                                                    wb.save("BookMyShowDatabase.xlsx")



                                    else:

                                        print('Eneter correct choice again')










                    break

                else:
                    flag = 1
            if flag == 1:
                print('You have entered an incorrect username or password')








elif rolePath == 2:
        print("******  Welcome to registration page **********")

        new_user     = input('Enter Username:  ')

        for i in range(1, ls_max_row + 1):
            # validating if user previously exists
            if login_sheet.cell(i, 1).value == new_user:
                flag = 1
                break
            else:
                # creating new cell to save username
                LoginData_User_row = ls_max_row + 1
                login_sheet.cell(row=LoginData_User_row, column=1, value=new_user)
        if flag == 1:
                print('User already exists')
                sys.exit()

        new_password = input('Enter Password:  ')
        # saving password on that particular cell
        login_sheet.cell(row=ls_max_row + 1, column=2, value=new_password)
        # enter user in Movie Rating sheet and maintained the same column
        movie_rating_sheet.cell(row=LoginData_User_row, column=1, value=new_user)
        # enter user in Movie Booking sheet and maintained the same column
        movie_booking_sheet.cell(row=LoginData_User_row, column=1, value=new_user)

        wb.save("BookMyShowDatabase.xlsx")

        if flag == 1:
            print('User already exists')

elif rolePath == 3:
    print('Thank you')


else:
    print('Please enter the correct choice')
    sys.exit()









